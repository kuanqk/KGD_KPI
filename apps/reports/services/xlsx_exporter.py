"""
XLSX-экспорт KPI-отчёта.

Структура файла:
  Лист 1 «Рейтинг»  — позиция, регион, K1-K6, итого (все 20 ДГД + КГД)
  Листы 2-7          — по одному KPI: регион, план, факт, % исполн., баллы
"""
import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from apps.kpi.models import KPIResult, KPISummary

# ── Стили ──────────────────────────────────────────────────────────────────

_HDR_FILL  = PatternFill('solid', fgColor='1F4E79')
_HDR_FONT  = Font(color='FFFFFF', bold=True)
_GRN_FILL  = PatternFill('solid', fgColor='D1FAE5')
_GRN_FONT  = Font(color='065F46', bold=True)
_YLW_FILL  = PatternFill('solid', fgColor='FEF3C7')
_YLW_FONT  = Font(color='92400E', bold=True)
_RED_FILL  = PatternFill('solid', fgColor='FEE2E2')
_RED_FONT  = Font(color='991B1B', bold=True)
_THIN      = Side(style='thin')
_BORDER    = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_CENTER    = Alignment(horizontal='center', vertical='center')

# ── Метаданные KPI ─────────────────────────────────────────────────────────

_KPI_META = [
    ('assessment',       'KPI 1 — Доначисление',        10, 'score_assessment'),
    ('collection',       'KPI 2 — Взыскание',            40, 'score_collection'),
    ('avg_assessment',   'KPI 3 — Среднее доначисление', 10, 'score_avg_assessment'),
    ('workload',         'KPI 4 — Коэф. занятости',      15, 'score_workload'),
    ('long_inspections', 'KPI 5 — Долгие проверки',      10, 'score_long_inspections'),
    ('cancelled',        'KPI 6 — Отменённые суммы',     15, 'score_cancelled'),
]

_SCORE_FIELDS = [m[3] for m in _KPI_META]


class XLSXExporter:
    def __init__(self, summary: KPISummary):
        self.summary = summary
        self.date_from = summary.date_from
        self.date_to = summary.date_to
        self.generated_at = datetime.now()

    # ── Публичный метод ────────────────────────────────────────────────────

    def generate(self) -> io.BytesIO:
        wb = Workbook()
        wb.remove(wb.active)  # убрать дефолтный лист

        all_summaries = list(
            KPISummary.objects
            .filter(date_from=self.date_from, date_to=self.date_to)
            .select_related('region')
            .order_by('rank', 'region__order')
        )
        dgd = [s for s in all_summaries if not s.region.is_summary]
        kgd = [s for s in all_summaries if s.region.is_summary]

        self._sheet_rating(wb, dgd, kgd)

        for kpi_type, label, max_score, _field in _KPI_META:
            results = list(
                KPIResult.objects
                .filter(kpi_type=kpi_type, date_from=self.date_from, date_to=self.date_to)
                .select_related('region')
                .exclude(region__is_summary=True)
                .order_by('region__order')
            )
            self._sheet_kpi(wb, label, max_score, results)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    # ── Приватные методы ───────────────────────────────────────────────────

    def _write_meta_rows(self, ws) -> int:
        """Записать 4-строчную шапку; вернуть номер строки для заголовка таблицы."""
        period = f'{self.date_from} — {self.date_to}'
        ws.cell(1, 1, 'KPI Monitor — КГД РК').font = Font(bold=True, size=13)
        ws.cell(2, 1, f'Период: {period}')
        ws.cell(3, 1, f'Дата формирования: {self.generated_at:%d.%m.%Y %H:%M}')
        return 5  # строка заголовка таблицы (строка 4 — пустой разделитель)

    def _hdr_cell(self, ws, row, col, value):
        c = ws.cell(row, col, value)
        c.fill   = _HDR_FILL
        c.font   = _HDR_FONT
        c.border = _BORDER
        c.alignment = _CENTER
        return c

    def _sheet_rating(self, wb, dgd, kgd):
        ws = wb.create_sheet('Рейтинг')
        hdr_row = self._write_meta_rows(ws)
        ws.cell(hdr_row - 1, 1, 'Сводный рейтинг ДГД').font = Font(bold=True, size=11)

        cols = ['#', 'Регион', 'K1', 'K2', 'K3', 'K4', 'K5', 'K6', 'Итого']
        for i, h in enumerate(cols, 1):
            self._hdr_cell(ws, hdr_row, i, h)

        ws.freeze_panes = ws.cell(hdr_row + 1, 1)

        data_row = hdr_row + 1
        for s in dgd:
            ws.cell(data_row, 1, s.rank or '').border = _BORDER
            ws.cell(data_row, 1).alignment = _CENTER
            ws.cell(data_row, 2, s.region.name_ru).border = _BORDER

            for col, field in enumerate(_SCORE_FIELDS, 3):
                c = ws.cell(data_row, col, getattr(s, field))
                c.border    = _BORDER
                c.alignment = _CENTER

            total = ws.cell(data_row, 9, s.score_total)
            total.border    = _BORDER
            total.alignment = _CENTER
            if s.score_total >= 80:
                total.fill, total.font = _GRN_FILL, _GRN_FONT
            elif s.score_total >= 50:
                total.fill, total.font = _YLW_FILL, _YLW_FONT
            else:
                total.fill, total.font = _RED_FILL, _RED_FONT

            data_row += 1

        if kgd:
            data_row += 1  # пустая строка-разделитель
            s = kgd[0]
            bold = Font(bold=True)
            ws.cell(data_row, 1, 'КГД').border = _BORDER
            ws.cell(data_row, 1).font = bold
            ws.cell(data_row, 2, s.region.name_ru).border = _BORDER
            ws.cell(data_row, 2).font = bold
            for col, field in enumerate(_SCORE_FIELDS, 3):
                c = ws.cell(data_row, col, getattr(s, field))
                c.border    = _BORDER
                c.alignment = _CENTER
                c.font      = bold
            c = ws.cell(data_row, 9, s.score_total)
            c.border    = _BORDER
            c.alignment = _CENTER
            c.font      = bold

        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 30
        for col_letter in ['C', 'D', 'E', 'F', 'G', 'H', 'I']:
            ws.column_dimensions[col_letter].width = 9

    def _sheet_kpi(self, wb, title: str, max_score: int, results):
        sheet_name = title[:31]
        ws = wb.create_sheet(sheet_name)
        hdr_row = self._write_meta_rows(ws)
        ws.cell(hdr_row - 1, 1, title).font = Font(bold=True, size=11)

        headers = ['Регион', 'План (млн тг)', 'Факт (млн тг)', '% исполнения', f'Баллы (макс. {max_score})']
        for i, h in enumerate(headers, 1):
            self._hdr_cell(ws, hdr_row, i, h)

        ws.freeze_panes = ws.cell(hdr_row + 1, 1)

        data_row = hdr_row + 1
        for r in results:
            ws.cell(data_row, 1, r.region.name_ru).border = _BORDER

            plan = round(float(r.plan) / 1_000_000, 4) if r.plan is not None else None
            fact = round(float(r.fact) / 1_000_000, 4) if r.fact is not None else None
            pct  = round(float(r.percent), 2)          if r.percent is not None else None

            c_plan = ws.cell(data_row, 2, plan)
            c_plan.border, c_plan.number_format = _BORDER, '#,##0.0000'

            c_fact = ws.cell(data_row, 3, fact)
            c_fact.border, c_fact.number_format = _BORDER, '#,##0.0000'

            c_pct = ws.cell(data_row, 4, pct)
            c_pct.border, c_pct.number_format = _BORDER, '0.00'
            c_pct.alignment = _CENTER

            c_score = ws.cell(data_row, 5, r.score)
            c_score.border, c_score.alignment = _BORDER, _CENTER

            data_row += 1

        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 16
        ws.column_dimensions['C'].width = 16
        ws.column_dimensions['D'].width = 14
        ws.column_dimensions['E'].width = 18
