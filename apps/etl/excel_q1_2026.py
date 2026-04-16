"""
Парсинг Excel «Статистика КЭР РК на 01.04.2026.xlsx» для тестовой загрузки Q1 2026.

Листы (ожидаемые имена): Доначисление, Взыскание, Среднее доначисление, Занятость,
Проводимые, Отмененные.
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

_CODE_RE = re.compile(r"^\d{2}xx$")


def _f(cell: Any) -> float | None:
    if cell is None:
        return None
    if isinstance(cell, (int, float)):
        return float(cell)
    try:
        return float(str(cell).replace(",", ".").replace(" ", ""))
    except (TypeError, ValueError):
        return None


def _i(cell: Any) -> int:
    if cell is None:
        return 0
    if isinstance(cell, int):
        return cell
    try:
        return int(float(str(cell).replace(",", ".")))
    except (TypeError, ValueError):
        return 0


@dataclass
class Q1RegionRow:
    """Одна строка ДГД из свода на 01.04.2026."""

    code: str
    # KPI 1 — Доначисление (млн тг)
    kbk_share_pct: float | None = None
    donachisleno_01_01_2026: float | None = None
    kpi_plan_01_04_2026: float | None = None
    kpi_fact_01_04_2026: float | None = None
    # KPI 2 — Взыскание (млн тг)
    vzyscano_01_01_2026: float | None = None
    k2_plan_01_04_2026: float | None = None
    k2_fact_01_04_2026: float | None = None
    # KPI 3 — Среднее доначисление
    check_count: int = 0
    donach_01_04_2025_mln: float | None = None
    donach_01_04_2026_mln: float | None = None
    # KPI 4 — Занятость
    staff_count: int = 0
    completed_checks_for_workload: int = 0
    # KPI 5 — Проводимые
    active_total: int = 0
    active_long: int = 0
    # KPI 6 — Отмененные (млн тг)
    assessed_total_mln: float | None = None
    cancelled_mln: float | None = None


@dataclass
class ParsedQ1Excel:
    path: Path
    regions: dict[str, Q1RegionRow] = field(default_factory=dict)


def _merge_region(dst: Q1RegionRow, **kwargs) -> None:
    for k, v in kwargs.items():
        if not hasattr(dst, k):
            continue
        if v is None:
            continue
        setattr(dst, k, v)


def parse_statistika_ker_2026_04(path: str | Path) -> ParsedQ1Excel:
    """
    Читает xlsx, собирает по коду ДГД (62xx) строки с листов KPI-таблиц.
    """
    path = Path(path)
    out = ParsedQ1Excel(path=path)
    wb = load_workbook(path, read_only=True, data_only=True)

    def get_sheet(names: tuple[str, ...]):
        for n in names:
            if n in wb.sheetnames:
                return wb[n]
        raise ValueError(f"Ни один из листов {names} не найден. Есть: {wb.sheetnames}")

    # ── Доначисление: A=код, D=доля КБК, E=на 01.01.2026, H=KPI на 01.04.2026, I=факт на 01.04.2026
    ws = get_sheet(("Доначисление", "Лист1"))
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=True):
        code = row[0] if row else None
        if not code or not isinstance(code, str):
            continue
        code = code.strip()
        if not _CODE_RE.match(code):
            continue
        r = out.regions.setdefault(code, Q1RegionRow(code=code))
        _merge_region(
            r,
            kbk_share_pct=_f(row[3]) if len(row) > 3 else None,
            donachisleno_01_01_2026=_f(row[4]) if len(row) > 4 else None,
            kpi_plan_01_04_2026=_f(row[7]) if len(row) > 7 else None,
            kpi_fact_01_04_2026=_f(row[8]) if len(row) > 8 else None,
        )

    # ── Взыскание: E=на 01.01.2026, H, I
    ws = get_sheet(("Взыскание", "Лист2"))
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=True):
        code = row[0] if row else None
        if not code or not isinstance(code, str):
            continue
        code = code.strip()
        if not _CODE_RE.match(code):
            continue
        r = out.regions.setdefault(code, Q1RegionRow(code=code))
        _merge_region(
            r,
            vzyscano_01_01_2026=_f(row[4]) if len(row) > 4 else None,
            k2_plan_01_04_2026=_f(row[7]) if len(row) > 7 else None,
            k2_fact_01_04_2026=_f(row[8]) if len(row) > 8 else None,
        )

    # ── Среднее доначисление: A=код, D=кол-во, F=на 01.04.2025, G=на 01.04.2026
    ws = get_sheet(("Среднее доначисление",))
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=True):
        code = row[0] if row else None
        if not code or not isinstance(code, str):
            continue
        code = code.strip()
        if not _CODE_RE.match(code):
            continue
        r = out.regions.setdefault(code, Q1RegionRow(code=code))
        _merge_region(
            r,
            check_count=_i(row[3]) if len(row) > 3 else 0,
            donach_01_04_2025_mln=_f(row[5]) if len(row) > 5 else None,
            donach_01_04_2026_mln=_f(row[6]) if len(row) > 6 else None,
        )

    # ── Занятость: A=код, D=штат, E=завершённые проверки
    ws = get_sheet(("Занятость",))
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=True):
        code = row[0] if row else None
        if not code or not isinstance(code, str):
            continue
        code = code.strip()
        if not _CODE_RE.match(code):
            continue
        r = out.regions.setdefault(code, Q1RegionRow(code=code))
        _merge_region(
            r,
            staff_count=_i(row[3]) if len(row) > 3 else 0,
            completed_checks_for_workload=_i(row[4]) if len(row) > 4 else 0,
        )

    # ── Проводимые: A=код, D=всего, E=долгих
    ws = get_sheet(("Проводимые",))
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True):
        code = row[0] if row else None
        if not code or not isinstance(code, str):
            continue
        code = code.strip()
        if not _CODE_RE.match(code):
            continue
        r = out.regions.setdefault(code, Q1RegionRow(code=code))
        _merge_region(
            r,
            active_total=_i(row[3]) if len(row) > 3 else 0,
            active_long=_i(row[4]) if len(row) > 4 else 0,
        )

    # ── Отмененные: A=код, D=всего доначислено, G=отменённая сумма (млн)
    ws = get_sheet(("Отмененные",))
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=True):
        code = row[0] if row else None
        if not code or not isinstance(code, str):
            continue
        code = code.strip()
        if not _CODE_RE.match(code):
            continue
        r = out.regions.setdefault(code, Q1RegionRow(code=code))
        assessed = _f(row[3]) if len(row) > 3 else None
        cancelled = _f(row[6]) if len(row) > 6 else None
        _merge_region(r, assessed_total_mln=assessed, cancelled_mln=cancelled)

    wb.close()
    return out
